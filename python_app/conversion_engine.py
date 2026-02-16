from __future__ import annotations

import io
import json
import math
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
MAPPING_FILE = BASE_DIR / "account_mapping.json"

with MAPPING_FILE.open("r", encoding="utf-8") as f:
    ACCOUNT_MAPPING: dict[str, dict[str, str]] = json.load(f)

BALANCE_GROUPS_TEMPLATE = {
    "Activo No Corriente": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Activo Corriente": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Patrimonio Neto": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Pasivo No Corriente": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Pasivo Corriente": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
}

PNL_SECTIONS_TEMPLATE = {
    "Importe neto cifra negocios": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Otros ingresos de explotacion": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Gastos de personal": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Servicios exteriores": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Tributos": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Amortizaciones": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Gastos excepcionales": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Resultado financiero": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
    "Otros resultados": {"items": [], "totalMXN": 0.0, "totalEUR": 0.0},
}


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "ñ": "n",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return "".join(ch for ch in text if ch.isalnum())


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        if math.isfinite(value):
            return float(value)
        return 0.0
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0.0
    text = text.replace(".", "").replace(",", ".")
    text = "".join(ch for ch in text if ch in "0123456789.-")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_zero_segment(segment: str) -> bool:
    s = str(segment or "").strip()
    return bool(s) and all(ch == "0" for ch in s)


def _normalize_row(row: dict[str, Any], index: int) -> dict[str, Any]:
    return {
        "_rowId": str(row.get("_rowId") or row.get("rowId") or row.get("id") or f"row-{index+1}"),
        "_isNew": bool(row.get("_isNew", False)),
        "_excludeFromAnalysis": bool(row.get("_excludeFromAnalysis", False)),
        "code": str(row.get("code", "")).strip(),
        "name": str(row.get("name", "Sin descripcion")).strip() or "Sin descripcion",
        "sid": _to_float(row.get("sid")),
        "sia": _to_float(row.get("sia")),
        "cargos": _to_float(row.get("cargos")),
        "abonos": _to_float(row.get("abonos")),
        "sfd": _to_float(row.get("sfd")),
        "sfa": _to_float(row.get("sfa")),
    }


def find_mapping(code: str) -> dict[str, str] | None:
    safe = str(code or "").strip()
    if not safe:
        return None
    parts = [p for p in safe.split("-") if p]
    candidates = [safe]
    for ln in range(len(parts), 0, -1):
        candidates.append("-".join(parts[:ln]))
    for c in candidates:
        if c in ACCOUNT_MAPPING:
            return ACCOUNT_MAPPING[c]
    return None


def _account_display_value(group: str, saldo: float) -> float:
    if group in {"Pasivo Corriente", "Pasivo No Corriente", "Patrimonio Neto", "Ingresos", "Ingresos Financieros"}:
        return -saldo
    return saldo


def _detect_summary_line(row: dict[str, Any], all_rows: list[dict[str, Any]]) -> bool:
    code = str(row.get("code", "")).strip()
    if not code:
        return False
    if code.startswith("000-000-"):
        return True

    segments = code.split("-")
    trailing_zero_count = 0
    for seg in reversed(segments):
        if _is_zero_segment(seg):
            trailing_zero_count += 1
        else:
            break

    if trailing_zero_count == 0:
        return False
    prefix_len = len(segments) - trailing_zero_count
    if prefix_len <= 0:
        return True
    prefix = "-".join(segments[:prefix_len]) + "-"
    return any(c.get("code") != code and str(c.get("code", "")).startswith(prefix) for c in all_rows)


def parse_workbook(file_bytes: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    matrix_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        matrix_rows.append(["" if v is None else v for v in row])

    header_row_idx = -1
    for idx, row in enumerate(matrix_rows):
        c0 = _norm_header(row[0] if len(row) > 0 else "")
        c1 = _norm_header(row[1] if len(row) > 1 else "")
        if c0 == "cuenta" and "nombre" in c1:
            header_row_idx = idx
            break

    rows: list[dict[str, Any]] = []
    if header_row_idx >= 0:
        start = header_row_idx + 1
        while start < len(matrix_rows):
            code = str(matrix_rows[start][0] if len(matrix_rows[start]) > 0 else "").strip()
            if code:
                break
            start += 1

        for raw in matrix_rows[start:]:
            code = str(raw[0] if len(raw) > 0 else "").strip()
            if not code or not any(ch.isdigit() for ch in code):
                continue
            rows.append(
                {
                    "_rowId": f"row-{len(rows)+1}",
                    "_isNew": False,
                    "_excludeFromAnalysis": False,
                    "code": code,
                    "name": str(raw[1] if len(raw) > 1 else "Sin descripcion").strip() or "Sin descripcion",
                    "sid": _to_float(raw[2] if len(raw) > 2 else 0),
                    "sia": _to_float(raw[3] if len(raw) > 3 else 0),
                    "cargos": _to_float(raw[4] if len(raw) > 4 else 0),
                    "abonos": _to_float(raw[5] if len(raw) > 5 else 0),
                    "sfd": _to_float(raw[6] if len(raw) > 6 else 0),
                    "sfa": _to_float(raw[7] if len(raw) > 7 else 0),
                }
            )

    if rows:
        return rows

    # fallback simple tabular scan
    df = pd.DataFrame(matrix_rows)
    out_rows: list[dict[str, Any]] = []
    for i in range(len(df)):
        row = df.iloc[i].tolist()
        code = str(row[0] if len(row) > 0 else "").strip()
        if not code:
            continue
        out_rows.append(
            {
                "_rowId": f"row-{len(out_rows)+1}",
                "_isNew": False,
                "_excludeFromAnalysis": False,
                "code": code,
                "name": str(row[1] if len(row) > 1 else "Sin descripcion").strip() or "Sin descripcion",
                "sid": _to_float(row[2] if len(row) > 2 else 0),
                "sia": _to_float(row[3] if len(row) > 3 else 0),
                "cargos": _to_float(row[4] if len(row) > 4 else 0),
                "abonos": _to_float(row[5] if len(row) > 5 else 0),
                "sfd": _to_float(row[6] if len(row) > 6 else 0),
                "sfa": _to_float(row[7] if len(row) > 7 else 0),
            }
        )
    return out_rows


def convert_rows(
    rows: list[dict[str, Any]],
    exchange_rate: float = 0.046,
    manual_mappings: dict[str, dict[str, str]] | None = None,
    period: dict[str, int] | None = None,
) -> dict[str, Any]:
    manual_mappings = manual_mappings or {}
    normalized_rows = [_normalize_row(r, i) for i, r in enumerate(rows) if str(r.get("code", "")).strip()]

    converted_data: list[dict[str, Any]] = []
    for row in normalized_rows:
        row_id = row["_rowId"]
        manual = manual_mappings.get(row_id, {})
        has_manual = any(str(manual.get(k, "")).strip() for k in ("pgc", "pgcName", "grupo", "subgrupo"))
        auto = find_mapping(row["code"])
        mapping = (
            {
                "pgc": str(manual.get("pgc", "")).strip() or "SIN MAPEO",
                "pgcName": str(manual.get("pgcName", "")).strip() or "Sin equivalencia PGC",
                "grupo": str(manual.get("grupo", "")).strip() or "Sin clasificar",
                "subgrupo": str(manual.get("subgrupo", "")).strip() or "Sin clasificar",
            }
            if has_manual
            else auto
        )

        saldo = row["sfd"] - row["sfa"]
        group = mapping["grupo"] if mapping else "Sin clasificar"
        display_mxn = _account_display_value(group, saldo)
        summary = _detect_summary_line(row, normalized_rows)
        exclude = bool(row.get("_excludeFromAnalysis", False) or summary)

        converted_data.append(
            {
                **row,
                "mapping": mapping,
                "pgcCode": mapping["pgc"] if mapping else "SIN MAPEO",
                "pgcName": mapping["pgcName"] if mapping else "Sin equivalencia PGC",
                "grupo": group,
                "subgrupo": mapping["subgrupo"] if mapping else "Sin clasificar",
                "saldo": saldo,
                "saldoEur": saldo * exchange_rate,
                "displayMXN": display_mxn,
                "displayEUR": display_mxn * exchange_rate,
                "manualMappingApplied": has_manual,
                "isSummaryLine": summary,
                "excludeFromAnalysis": exclude,
            }
        )

    rows_for_analysis = [r for r in converted_data if not r["excludeFromAnalysis"]]

    aggregate_map: dict[str, dict[str, Any]] = {}
    for row in rows_for_analysis:
        key = row["pgcCode"]
        if key not in aggregate_map:
            aggregate_map[key] = {
                "pgcCode": row["pgcCode"],
                "pgcName": row["pgcName"],
                "grupo": row["grupo"],
                "subgrupo": row["subgrupo"],
                "totalMXN": 0.0,
                "totalEUR": 0.0,
                "details": [],
            }
        aggregate_map[key]["totalMXN"] += row["displayMXN"]
        aggregate_map[key]["totalEUR"] += row["displayEUR"]
        aggregate_map[key]["details"].append(row)

    pgc_aggregated = sorted(aggregate_map.values(), key=lambda x: str(x["pgcCode"]))

    balance_groups = deepcopy(BALANCE_GROUPS_TEMPLATE)
    for row in pgc_aggregated:
        if row["grupo"] not in balance_groups:
            continue
        balance_groups[row["grupo"]]["items"].append(row)
        balance_groups[row["grupo"]]["totalMXN"] += row["totalMXN"]
        balance_groups[row["grupo"]]["totalEUR"] += row["totalEUR"]

    total_activo_mxn = balance_groups["Activo No Corriente"]["totalMXN"] + balance_groups["Activo Corriente"]["totalMXN"]
    total_pasivo_pn_mxn = (
        balance_groups["Patrimonio Neto"]["totalMXN"]
        + balance_groups["Pasivo No Corriente"]["totalMXN"]
        + balance_groups["Pasivo Corriente"]["totalMXN"]
    )
    total_activo_eur = balance_groups["Activo No Corriente"]["totalEUR"] + balance_groups["Activo Corriente"]["totalEUR"]
    total_pasivo_pn_eur = (
        balance_groups["Patrimonio Neto"]["totalEUR"]
        + balance_groups["Pasivo No Corriente"]["totalEUR"]
        + balance_groups["Pasivo Corriente"]["totalEUR"]
    )

    diff_mxn = total_activo_mxn - total_pasivo_pn_mxn
    diff_eur = total_activo_eur - total_pasivo_pn_eur

    adjusted_total_pasivo_pn_mxn = total_pasivo_pn_mxn
    adjusted_total_pasivo_pn_eur = total_pasivo_pn_eur
    auto_result_line = None
    if abs(diff_mxn) > 0.01:
        auto_result_line = {
            "pgcCode": "129",
            "pgcName": "Resultado del periodo pendiente de cierre",
            "grupo": "Patrimonio Neto",
            "subgrupo": "Fondos propios",
            "totalMXN": diff_mxn,
            "totalEUR": diff_eur,
            "details": [],
        }
        adjusted_total_pasivo_pn_mxn += diff_mxn
        adjusted_total_pasivo_pn_eur += diff_eur

    pnl_sections = deepcopy(PNL_SECTIONS_TEMPLATE)
    for row in pgc_aggregated:
        sub = row["subgrupo"]
        if sub in pnl_sections:
            pnl_sections[sub]["items"].append(row)
            pnl_sections[sub]["totalMXN"] += row["totalMXN"]
            pnl_sections[sub]["totalEUR"] += row["totalEUR"]

    ingresos_mx = pnl_sections["Importe neto cifra negocios"]["totalMXN"] + pnl_sections["Otros ingresos de explotacion"]["totalMXN"]
    gastos_mx = (
        pnl_sections["Gastos de personal"]["totalMXN"]
        + pnl_sections["Servicios exteriores"]["totalMXN"]
        + pnl_sections["Tributos"]["totalMXN"]
        + pnl_sections["Amortizaciones"]["totalMXN"]
        + pnl_sections["Gastos excepcionales"]["totalMXN"]
    )
    resultado_fin_mx = pnl_sections["Resultado financiero"]["totalMXN"]
    otros_res_mx = pnl_sections["Otros resultados"]["totalMXN"]
    resultado_antes_mx = ingresos_mx - gastos_mx + resultado_fin_mx + otros_res_mx

    ingresos_eur = pnl_sections["Importe neto cifra negocios"]["totalEUR"] + pnl_sections["Otros ingresos de explotacion"]["totalEUR"]
    gastos_eur = (
        pnl_sections["Gastos de personal"]["totalEUR"]
        + pnl_sections["Servicios exteriores"]["totalEUR"]
        + pnl_sections["Tributos"]["totalEUR"]
        + pnl_sections["Amortizaciones"]["totalEUR"]
        + pnl_sections["Gastos excepcionales"]["totalEUR"]
    )
    resultado_fin_eur = pnl_sections["Resultado financiero"]["totalEUR"]
    otros_res_eur = pnl_sections["Otros resultados"]["totalEUR"]
    resultado_antes_eur = ingresos_eur - gastos_eur + resultado_fin_eur + otros_res_eur

    unmapped_rows = [r for r in rows_for_analysis if r["pgcCode"] == "SIN MAPEO"]

    total_debe_inicial = sum(r["sid"] for r in rows_for_analysis)
    total_haber_inicial = sum(r["sia"] for r in rows_for_analysis)
    total_debe_final = sum(r["sfd"] for r in rows_for_analysis)
    total_haber_final = sum(r["sfa"] for r in rows_for_analysis)

    return {
        "metadata": {
            "exchangeRate": exchange_rate,
            "rowCount": len(converted_data),
            "analyzedRowCount": len(rows_for_analysis),
            "summaryExcludedCount": sum(1 for r in converted_data if r["isSummaryLine"]),
            "unmappedCount": len(unmapped_rows),
            "manualMappingCount": sum(1 for r in converted_data if r["manualMappingApplied"]),
            "mappedCoveragePct": ((len(rows_for_analysis) - len(unmapped_rows)) / len(rows_for_analysis) * 100) if rows_for_analysis else 0.0,
            "period": period,
        },
        "convertedData": converted_data,
        "pgcAggregated": pgc_aggregated,
        "balanceSheet": {
            "groups": balance_groups,
            "totalActivoMXN": total_activo_mxn,
            "totalPasivoPNMXN": total_pasivo_pn_mxn,
            "totalActivoEUR": total_activo_eur,
            "totalPasivoPNEUR": total_pasivo_pn_eur,
            "differenceMXN": diff_mxn,
            "differenceEUR": diff_eur,
            "autoResultLine": auto_result_line,
            "adjustedTotalPasivoPNMXN": adjusted_total_pasivo_pn_mxn,
            "adjustedTotalPasivoPNEUR": adjusted_total_pasivo_pn_eur,
            "adjustedDifferenceMXN": total_activo_mxn - adjusted_total_pasivo_pn_mxn,
            "adjustedDifferenceEUR": total_activo_eur - adjusted_total_pasivo_pn_eur,
        },
        "pnl": {
            "sections": pnl_sections,
            "ingresosMx": ingresos_mx,
            "gastosMx": gastos_mx,
            "resultadoExplotacionMx": ingresos_mx - gastos_mx,
            "resultadoFinancieroMx": resultado_fin_mx,
            "otrosResultadosMx": otros_res_mx,
            "resultadoAntesImpuestosMx": resultado_antes_mx,
            "ingresosEur": ingresos_eur,
            "gastosEur": gastos_eur,
            "resultadoExplotacionEur": ingresos_eur - gastos_eur,
            "resultadoFinancieroEur": resultado_fin_eur,
            "otrosResultadosEur": otros_res_eur,
            "resultadoAntesImpuestosEur": resultado_antes_eur,
        },
        "validations": {
            "trialBalanceInitialDifference": total_debe_inicial - total_haber_inicial,
            "trialBalanceFinalDifference": total_debe_final - total_haber_final,
            "unmappedRows": unmapped_rows,
        },
    }


def export_conversion_xlsx(conversion: dict[str, Any]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        pd.DataFrame(conversion["convertedData"]).to_excel(writer, index=False, sheet_name="Mapeo_Detalle")
        pd.DataFrame(conversion["pgcAggregated"]).to_excel(writer, index=False, sheet_name="Balanza_PGC")
        pd.DataFrame([
            {"Control": "Lineas totales", "Valor": conversion["metadata"]["rowCount"]},
            {"Control": "Lineas analizadas", "Valor": conversion["metadata"]["analyzedRowCount"]},
            {"Control": "Sin mapear", "Valor": conversion["metadata"]["unmappedCount"]},
            {"Control": "Cobertura %", "Valor": conversion["metadata"]["mappedCoveragePct"]},
            {"Control": "Dif. balanza final", "Valor": conversion["validations"]["trialBalanceFinalDifference"]},
        ]).to_excel(writer, index=False, sheet_name="Validaciones")
    return output.getvalue()
